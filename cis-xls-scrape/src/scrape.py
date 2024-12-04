from hashlib import sha256
from importlib.metadata import distribution
import openpyxl
import os
import sys
from dataclasses import dataclass, field, asdict
from typing import List, Optional
import yaml
from enum import Enum, unique
import input_definitions
from collections import deque
import codegen

class ProcedureType(Enum):
    AUDIT = 1
    REMEDIATION = 2

@dataclass
class OpenAPISpec:
    openapi: str = "3.0.0"
    info: dict = field(default_factory=lambda: {
        "title": "CIS for Linux osconfig module API",
        "description": "Aim of this API is to provide visualisation for the CIS for Linux osconfig module recommendations",
        "version": "1.0.0"
    })
    paths: dict = field(default_factory=dict)
    components: dict = field(default_factory=dict)
    tags: List[dict] = field(default_factory=list)
    servers: List[dict] = "http://localhost:8080"

@dataclass
class BashScriptSchema:
    type: str = "string"
    title: str = "Bash script content"
    description: str = "The content of the bash script"
    example: str = "echo 'Hello, World!'"

@dataclass
class Recommendation:
    path: str
    title: str
    description: str
    rationale: str
    remediation_procedure: Optional[str] = None
    audit_procedure: Optional[str] = None

@dataclass
class RecommendationNumberSchema:
    type: str = "string"
    title: str = "Recommendation number"
    description: str = "CIS recommendation numbers are dot-separated numbers that uniquely identify a recommendation"
    example: str = "1.1.1"

@dataclass
class RecommendationTitleSchema:
    type: str = "string"
    title: str = "Recommendation title"
    description: str = "The title of the recommendation"
    example: str = "Ensure mounting of cramfs filesystems is disabled"

@dataclass
class RecommendationDescriptionSchema:
    type: str = "string"
    title: str = "Recommendation description"
    description: str = "The description of the recommendation"
    example: str = "The cramfs filesystem type is a compressed read-only Linux filesystem embedded in small footprint systems. A cramfs image can be used without having to first decompress the image. The cramfs image is read-only, and cannot be modified. The cramfs filesystem type is being phased out in favor of squashfs, which provides greater flexibility and the ability to be mounted read-write. The cramfs filesystem type should be disabled unless required by your organization."

@dataclass
class RecommendationRationaleSchema:
    type: str = "string"
    title: str = "Recommendation rationale"
    description: str = "The rationale behind the recommendation"
    example: str = "The cramfs filesystem type is a compressed read-only Linux filesystem embedded in small footprint systems. A cramfs image can be used without having to first decompress the image. The cramfs image is read-only, and cannot be modified. The cramfs filesystem type is being phased out in favor of squashfs, which provides greater flexibility and the ability to be mounted read-write. The cramfs filesystem type should be disabled unless required by your organization."


@dataclass
class RecommendationSchema:
    type: str = "object"
    properties: dict = field(default_factory=lambda: {
        "number": {"$ref": "#/components/schemas/RecommendationNumberSchema"},
        "title": {"$ref": "#/components/schemas/RecommendationTitleSchema"},
        "description": {"$ref": "#/components/schemas/RecommendationDescriptionSchema"},
        "rationale": {"$ref": "#/components/schemas/RecommendationRationaleSchema"},
        "remediation_procedure": {"$ref": "#/components/schemas/BashScript"},
        "audit_procedure": {"$ref": "#/components/schemas/BashScript"}
    })
    required: List[str] = field(default_factory=lambda: ["number", "title", "description", "rationale"])

@dataclass
class Tag:
    name: str
    description: str

    def __hash__(self):
        return hash(self.name)

@dataclass
class Method:
    summary: str
    operationId: str
    tags: List[str]
    # requestBody: dict
    responses: dict

tags = {
    Tag(name="Audit", description="Audit recommendations"),
    Tag(name="Remediation", description="Remediation recommendations"),
    Tag(name="CIS", description="Center for Internet Security"),
    Tag(name="CIS-L1", description="CIS Level 1"),
    Tag(name="CIS-L1-Server", description="CIS Level 1 Server"),
    Tag(name="CIS-L1-Workstation", description="CIS Level 1 Server"),
    Tag(name="CIS-L2-Server", description="CIS Level 2 Server"),
    Tag(name="CIS-L2-Workstation", description="CIS Level 2 Server")
    }
paths = {}

api = OpenAPISpec()
api.components['schemas'] = {
    "RecommendationNumberSchema": RecommendationNumberSchema(),
    "RecommendationTitleSchema": RecommendationTitleSchema(),
    "RecommendationDescriptionSchema": RecommendationDescriptionSchema(),
    "RecommendationRationaleSchema": RecommendationRationaleSchema(),
    "Recommendation": RecommendationSchema(),
    "BashScript": BashScriptSchema()
}
api.tags = tags

@dataclass
class Stats:
    rules: set[str]

stats = dict[Stats]()

all_rules = set[str]()
unique_rules = set[str]()
titles = list[str]()

output_directory = None


def make_audit_method(title:str, number: str, category: str, machine_type: str, input_definition, procedure_hash: str):
    return Method(
        summary=title,
        operationId=procedure_hash,
        tags=["Audit", "CIS", f"CIS-{category}", f"CIS-{category}-{machine_type}", input_definition.distribution_name, f"{input_definition.distribution_name}-{input_definition.distribution_version}"],
        # requestBody={"$ref": "#/components/schemas/Recommendation"},
        responses={
            "200": {
                "description": "Compliant"
            },
            "400": {
                "description": "Non-compliant"
            },
            "500": {
                "description": "Audit was not performed due to an error"
            }
        }
    )




def make_remediation_method(title:str, number: str, category: str, machine_type: str, input_definition, procedure_hash: str):
    return Method(
        summary=title,
        operationId=procedure_hash,
        tags=["Remediation", "CIS", f"CIS-{category}", f"CIS-{category}-{machine_type}", input_definition.distribution_name, f"{input_definition.distribution_name}-{input_definition.distribution_version}"],
        # requestBody={"$ref": "#/components/schemas/Recommendation"},
        responses={
            "200": {
                "description": "Remediation succeeded"
            },
            "500": {
                "description": "Remediation was not performed due to an error"
            }
        }
    )



def update_method_tags(api_path: str, method: str, category: str, machine_type: str, input_definition):
    existing_tags = {tag for tag in paths[api_path][method].tags}
    existing_tags.add("CIS")
    existing_tags.add(f"CIS-{category}")
    existing_tags.add(f"CIS-{category}-{machine_type}")
    existing_tags.add(input_definition.distribution_name)
    existing_tags.add(f"{input_definition.distribution_name}-{input_definition.distribution_version}")
    paths[api_path][method].tags = list(existing_tags)




def load_recommendation(row, category: str, machine_type: str, input_definition, procedure_type: ProcedureType):
    recommendation_number = row[input_definition.recommendation_idx]
    recommendation_path = row[input_definition.recommendation_idx].replace('.', '/')
    title = row[input_definition.title_idx]

    # Create the recommendation subdirectory path
    subdirectory_path = os.path.join(output_directory, "benchmarks", input_definition.benchmark_name, input_definition.distribution_name, input_definition.distribution_version, input_definition.benchmark_version, recommendation_path)

    # Recursively create the subdirectory
    os.makedirs(subdirectory_path, exist_ok=True)

    api_path = f"/benchmarks/{input_definition.benchmark_name}/{input_definition.distribution_name}/{input_definition.distribution_version}/{input_definition.benchmark_version}/{recommendation_path}"
    # print(f"{recommendation_number} {distribution} {version}")
    if api_path not in paths:
        paths[api_path] = {}

    procedure_idx = None
    # procedure_file_path = None
    if procedure_type == ProcedureType.AUDIT:
        procedure_idx = input_definition.audit_procedure_idx
        # procedure_file_path = os.path.join(subdirectory_path, 'audit_procedure.sh')
    elif procedure_type == ProcedureType.REMEDIATION:
        procedure_idx = input_definition.remediation_procedure_idx
        # procedure_file_path = os.path.join(subdirectory_path, 'remediation_procedure.sh')

    if not row[procedure_idx]:
        return

    # Extract text between ``` delimiters in the procedure field
    procedure = row[procedure_idx]
    script_start = procedure.find('```') + 3
    script_end = procedure.find('```', script_start)
    if script_start == -1 or script_end == -1:
        # print(f"Note: Missing remediation script for {recommendation_number} ({title}).")
        return

    procedure_content = procedure[script_start:script_end].strip()
    procedure_hash = sha256(procedure_content.encode()).hexdigest()
    procedure_path = os.path.join(output_directory, 'actions', f"{procedure_hash}.sh")

    # Write the extracted procedure to a file
    if not os.path.isfile(procedure_path):
        with open(procedure_path, 'w') as f:
            f.write(procedure_content)

    if input_definition.distribution_version not in stats:
        stats[input_definition.distribution_version] = Stats(set[str]())
    stats[input_definition.distribution_version].rules.add(recommendation_number)
    if recommendation_number in all_rules and recommendation_number in unique_rules:
        unique_rules.remove(recommendation_number)
    elif recommendation_number not in all_rules:
        unique_rules.add(recommendation_number)
    all_rules.add(recommendation_number)

    if procedure_type == ProcedureType.AUDIT:
        padded_title = title.ljust(80)
        titles.append(f"{padded_title}{input_definition.distribution_name} {input_definition.distribution_version}")
        procedure_path = os.path.join(subdirectory_path, 'audit.sh')
        if not os.path.isfile(procedure_path):
            with open(procedure_path, 'w') as f:
                f.write(procedure_content)
        method = make_audit_method(title, recommendation_number, category, machine_type, input_definition, procedure_hash)
        if 'get' not in paths[api_path]:
            paths[api_path]['get'] = method
        else:
            update_method_tags(api_path, 'get', category, machine_type, input_definition)
    if procedure_type == ProcedureType.REMEDIATION:
        procedure_path = os.path.join(subdirectory_path, 'remediation.sh')
        if not os.path.isfile(procedure_path):
            with open(procedure_path, 'w') as f:
                f.write(procedure_content)
        method = make_remediation_method(title, recommendation_number, category, machine_type, input_definition, procedure_hash)
        if 'post' not in paths[api_path]:
            paths[api_path]['post'] = method
        else:
            update_method_tags(api_path, 'post', category, machine_type, input_definition)




def load_sheet_data(sheet, category: str, machine_type: str, distribution: str, version: str):
    distro_found = False
    distro_version_found = False
    for tag in tags:
        if tag.name == distribution:
            distro_found = True
        if tag.name == f"{distribution}-{version}":
            distro_version_found = True

    if not distro_found:
        tags.add(Tag(distribution, f"CIS recommendations for {distribution} distribution"))
    if not distro_version_found:
        tags.add(Tag(f"{distribution}-{version}", f"CIS recommendations for {distribution} {version}"))

    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Skip rows without recommendation number
        if row[input_definition.recommendation_idx] is None:
            continue
        # Skip rows without 'Automated' assessment status
        # if row[input_definition.assesment_status_idx] != 'Automated':
        #     continue

        load_recommendation(row, category, machine_type, input_definition, ProcedureType.AUDIT)
        load_recommendation(row, category, machine_type, input_definition, ProcedureType.REMEDIATION)




if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python scrape.py <input_file_path> <output_directory>")
        sys.exit(1)

    input_directory = sys.argv[1]
    output_directory = sys.argv[2]

    if not os.path.isdir(input_directory):
        print(f"Error: {input_directory} must be a directory.")
        sys.exit(1)

    if not os.path.isdir(output_directory):
        print(f"Error: {output_directory} must be a directory.")
        sys.exit(1)

    os.makedirs(os.path.join(output_directory, 'actions'), exist_ok=True)

    for input_definition in input_definitions.get_input_definitions():
        if input_definition.benchmark_name != "cis":
            print(f"Error: Unsupported benchmark name: {input_definition.benchmark_name}")
            sys.exit(1)

        print(f"Processing file: {os.path.join(input_directory, input_definition.file_name)}")

        # List all sheet names in the workbook
        workbook = openpyxl.load_workbook(os.path.join(input_directory, input_definition.file_name))
        sheet_names = workbook.sheetnames
        # print("Available sheets:")
        # for sheet_name in sheet_names:
        #     print(f"- {sheet_name}")

        # Load data from the specific sheets 'Level 1 - Server'
        if 'Level 1 - Server' in sheet_names:
            load_sheet_data(workbook["Level 1 - Server"], "L1", "Server", input_definition.distribution_name, input_definition.distribution_version)

        # if 'Level 1 - Workstation' in sheet_names:
        #     load_sheet_data(workbook["Level 1 - Workstation"], "L1", "Workstation", input_definition.distribution_name, input_definition.distribution_version)

        # if 'Level 2 - Server' in sheet_names:
        #     load_sheet_data(workbook["Level 2 - Server"], "L2", "Server", input_definition.distribution_name, input_definition.distribution_version)

        # if 'Level 2 - Workstation' in sheet_names:
        #     load_sheet_data(workbook["Level 2 - Workstation"], "L2", "Workstation", input_definition.distribution_name, input_definition.distribution_version)

        if 'Level 1' in sheet_names:
            load_sheet_data(workbook["Level 1"], "L1", "Server", input_definition.distribution_name, input_definition.distribution_version)

        # if 'Level 2' in sheet_names:
        #     load_sheet_data(workbook["Level 2"], "L2", "Server", input_definition.distribution_name, input_definition.distribution_version)

    api.paths = paths
    api.tags = list(tags)

    # Save the OpenAPI specification to a yaml file
    openapi_file_path = os.path.join(output_directory, 'openapi.yaml')
    with open(openapi_file_path, 'w') as f:
        yaml.dump(asdict(api), f)

    mof_file_path = os.path.join(output_directory, f"{input_definition.benchmark_name}.mof")
    with open(mof_file_path, 'w') as mof_file:
        for path in api.paths:
            if 'get' not in api.paths[path] or 'post' not in api.paths[path]:
                continue
            get = api.paths[path]['get']
            post = api.paths[path]['post']
            mof_file.write("instance of OsConfigResource {\n")
            mof_file.write(f'    ResourceID = "{get.summary}";\n')
            mof_file.write(f'    PayloadKey = "{path}";\n')
            mof_file.write(f'    ComponentName = "CIS_for_Linux";\n')
            mof_file.write(f'    InitObjectName = "{path}/init";\n')
            mof_file.write(f'    ReportedObjectName = "{path}/audit";\n')
            mof_file.write(f'    ExpectedObjectValue = "PASS";\n')
            mof_file.write(f'    DesiredObjectName = "{path}/remediate";\n')
            mof_file.write(f'    DesiredObjectValue = "";\n')
            mof_file.write(f'    ModuleName = "GuestConfiguration";\n')
            mof_file.write(f'    ModuleVersion = "1.0.0";\n')
            mof_file.write(f'    ConfigurationName = "CIS_for_Linux";\n')
            mof_file.write(f'    SourceInfo = "::4::5::OsConfigResource";\n')
            mof_file.write("};\n\n")

    codegen.generate_code(paths, output_directory, 'cis')

    csv_file_path = os.path.join(output_directory, f"{input_definition.benchmark_name}.csv")
    with open(csv_file_path, 'w') as csv_file:
        for path in api.paths:
            if 'get' not in api.paths[path] or 'post' not in api.paths[path]:
                continue
            split_path = path.split('/')
            distribution = split_path[3]
            version = split_path[4]
            cis_version = split_path[5]
            csv_file.write(f"{distribution},{version},{cis_version},{'.'.join(split_path[6:])},{api.paths[path]['get'].summary.replace(',','?')}\n")

    # Store paths in buckets based on the distribution and CIS version
    buckets = dict()
    for path in api.paths:
        split_path = path.split('/')
        bucket_name = '/'.join(split_path[2:6])
        if bucket_name not in buckets:
            buckets[bucket_name] = list[str]()
        buckets[bucket_name].append(path)

    # Print the buckets
    for bucket_name, bucket in buckets.items():
        os.makedirs(os.path.join(output_directory, 'mof', bucket_name), exist_ok=True)
        output_path = os.path.join(output_directory, 'mof', bucket_name, f"{bucket_name.replace('/', '.')}.mof")
        print(f"Bucket: {bucket_name} -> {output_path}")
        with open(output_path, 'w') as mof_file:
            for path in bucket:
                if 'get' not in api.paths[path] or 'post' not in api.paths[path]:
                    continue
                get = api.paths[path]['get']
                post = api.paths[path]['post']
                mof_file.write("instance of OsConfigResource {\n")
                mof_file.write(f'    ResourceID = "{get.summary}";\n')
                mof_file.write(f'    PayloadKey = "{path}";\n')
                mof_file.write(f'    ComponentName = "CIS_for_Linux";\n')
                mof_file.write(f'    InitObjectName = "{path}/init";\n')
                mof_file.write(f'    ReportedObjectName = "{path}/audit";\n')
                mof_file.write(f'    ExpectedObjectValue = "PASS";\n')
                mof_file.write(f'    DesiredObjectName = "{path}/remediate";\n')
                mof_file.write(f'    DesiredObjectValue = "";\n')
                mof_file.write(f'    ModuleName = "GuestConfiguration";\n')
                mof_file.write(f'    ModuleVersion = "1.0.0";\n')
                mof_file.write(f'    ConfigurationName = "CIS_for_Linux";\n')
                mof_file.write(f'    SourceInfo = "::4::5::OsConfigResource";\n')
                mof_file.write("};\n\n")
            # print(f"- {path}")
    # # Print statistics
    # print("Statistics:")
    # print(f"Total recommentations: {len(all_rules)}")
    # print(f"Total unique recommendations: {len(unique_rules)}")
    # for version, stat in stats.items():
    #     print(f"Version {version}: {len(stat.rules)} recommendations")
    # i = 0
    # for rule in sorted(list(all_rules)):
    #     i += 1
    #     print(f"Rule[{i}]: {rule}")
    # for title in sorted(titles):
    #     print(title)
