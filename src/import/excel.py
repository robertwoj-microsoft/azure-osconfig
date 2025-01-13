from enum import Enum
from math import perm
from os import path, sep
from struct import pack
from tokenize import group
from typing import List
import openpyxl
import sys
import input_definitions
from model import CISBenchmark, FunctionArgument, Recommendation, CISModel, FunctionCall, OrExpression, AndExpression, NotExpression
import re

class CISLevel(Enum):
    L1 = 1
    L2 = 2

class MachineType(Enum):
    Server = 1
    Workstation = 2

RECOMMENDATION_IDX = 1      # Column B
SUMMARY_IDX = 2             # Column C
ASSESMENT_STATUS_IDX = 3    # Column D
DESCRIPTION_IDX = 4         # Column E
AUDIT_IDX = 8               # Column I
REMEDIATION_IDX = 7         # Column H
# DEFAULT_VALUES_IDX = 24     # Column Y

def load_procedure(row, procedure_idx) -> str:
    if not row[procedure_idx]:
        return None
    procedure = row[procedure_idx]
    begin = procedure.find('```') + 3
    end = procedure.find('```', begin)
    if begin == -1 or end == -1:
        return None
    return procedure[begin:end].strip()

def load_procedures(row, procedure_idx) -> List[str]:
    if not row[procedure_idx]:
        return None
    result = []
    procedure = row[procedure_idx]
    begin = end = 0
    while begin != -1 and end != -1:
        begin = procedure.find('```', end) + 3
        end = procedure.find('```', begin)
        if begin == -1 or end == -1:
            break
        result.append(procedure[begin:end].strip())
    return result

def load_stat_command_output(procedure) -> str:
    begin = procedure.find('\n\n')
    if begin == -1:
        return None
    begin += 2

    procedure = procedure[begin:]
    end = procedure.find('\n')
    if end == -1:
        return procedure.strip()
    return procedure[:end].strip()

def generic_ensure_file_permissions(row, match, action: str) -> tuple:
    filename = match.group(1)

    procedure = load_procedure(row, AUDIT_IDX)
    if not procedure:
        return None

    output = None
    if row[AUDIT_IDX].startswith('Run the following command'):
        if procedure.startswith('# stat'):
            output = load_stat_command_output(procedure)
        elif procedure.startswith(f"# [ -e \"{filename}\" ] && stat") or procedure.startswith(f"# [ -e {filename} ] && stat"):
            output = load_stat_command_output(procedure)
    elif row[AUDIT_IDX].startswith('**- IF -** cron is installed on the system'):
        output = load_stat_command_output(procedure)

    if not output:
        print(f"Failed to load stat command output for {row[RECOMMENDATION_IDX]} - {row[SUMMARY_IDX]}")
        # print(procedure)
        return None

    # default_value = row[DEFAULT_VALUES_IDX]
    # if default_value == "File doesn't exist":
    #     return {
    #         "ensureFileDoesNotExist": {
    #             "filename": filename,
    #         }
    #     }

    # if default_value is None:
    #     return None

    pattern = r"/[^ ]+ (\d+) \d+/(\w+) \d+/(\w+)"
    match = re.match(pattern, output)
    if match:
        return (
            match.group(1),
            {
                "ensureFilePermissions": {
                    "filename": filename,
                    "user": match.group(2),
                    "group": match.group(3),
                    "permissions": match.group(1),
                }
            }
        )

    pattern = r"Access: \(([01234567]+)/[drwx-]{10}\) Uid: \( \d+/ (\w+)\) Gid: [\(\{] \d+/ (\w+)\)"
    match = re.match(pattern, output)
    if match:
        return (
            match.group(1),
            {
                "ensureFilePermissions": {
                    "filename": filename,
                    "user": match.group(2),
                    "group": match.group(3),
                    "permissions": match.group(1),
                }
            }
        )

    pattern = r"/[^ ]+ Access: \(([01234567]+)/[drwx-]{10}\) Uid: \( \d+/ (\w+)\) Gid: [\(\{] \d+/ (\w+)\)"
    match = re.match(pattern, output)
    if match:
        return (
            match.group(1),
            {
                "ensureFilePermissions": {
                    "filename": filename,
                    "user": match.group(2),
                    "group": match.group(3),
                    "permissions": match.group(1),
                }
            }
        )

    print(f"Error: Invalid output format: {output}")
    return None

def generic_ensure_partition_options(path: str, option: str, action: str) -> tuple:
    return (
        None,
        {
            "ensurePartitionOptionIsSet": {
                "path": path,
                "option": option,
            }
        }
    )

def generic_ensure_partition_options1(row, match, action: str) -> tuple:
    return generic_ensure_partition_options(match.group(2), match.group(1), action)

def generic_ensure_partition_options2(row, match, action: str) -> tuple:
    return generic_ensure_partition_options(match.group(1), match.group(2), action)

def generic_ensure_mounting_filesystems_disabled(row, match, action: str) -> tuple:
    return (
        None,
        {
            "ensureMountingOfFilesytemsIsDisabled": {
                "filesystem": match.group(1),
            }
        }
    )

def generic_ensure_service_not_in_use(row, match, action: str) -> tuple:
    activeServices = []
    enabledServices = []
    installedPackage = None
    procedures = load_procedures(row, AUDIT_IDX)
    for procedure in procedures:
        prefix = '# systemctl is-active '
        if procedure.startswith(prefix):
            begin = len(prefix)
            while True:
                end = procedure.find('.service', begin)
                if end == -1:
                    break
                activeServices.append(procedure[begin:end].strip())
                begin = end + len('.service')
            continue

        prefix = '# systemctl is-enabled '
        if procedure.startswith(prefix):
            begin = len(prefix)
            while True:
                end = procedure.find('.service', begin)
                if end == -1:
                    break
                enabledServices.append(procedure[begin:end].strip())
                begin = end + len('.service')
            continue

        # RHEL-based systems package check
        prefix = '# rpm -q '
        if procedure.startswith(prefix):
            procedure = procedure[len(prefix):]
            end = procedure.find('\n')
            if end != -1:
                procedure = procedure[:end]
            end = procedure.find(' ')
            if end != -1:
                procedure = procedure[:end]
            installedPackage = procedure.strip()
            continue

        # Debian-based systems package check
        prefix = '# dpkg-query -s '
        if procedure.startswith(prefix):
            procedure = procedure[len(prefix):]
            end = procedure.find('\n')
            if end != -1:
                procedure = procedure[:end]
            end = procedure.find(' ')
            if end != -1:
                procedure = procedure[:end]
            installedPackage = procedure.strip()
            continue

    # fallback
    if len(activeServices) == 0 or len(enabledServices) == 0 or installedPackage is None:
        print("Unable to obtain service information")
        return None

    serviceDisabledRules = []
    for service in enabledServices:
        rule = {
            "ensureServiceIsDisabled": {
                "service": service,
            }
        }
        serviceDisabledRules.append(rule)

    serviceInactiveRules = []
    for service in activeServices:
        rule = {
            "ensureServiceIsInactive": {
                "service": service,
            }
        }
        serviceInactiveRules.append(rule)

    packageNotInstalledRule = {
        "packageNotInstalled": {
            "name": installedPackage,
        }
    }

    orRule = {
        "OR": [
            packageNotInstalledRule,
            {
                "AND": serviceDisabledRules
            }
        ]
    }

    # Option 3 output
    return (
        None,
        {
            "AND": serviceInactiveRules + [orRule]
        }
    )

    # Option 2 output
    # return (
    #     None,
    #     {
    #         "ensureServiceIsNotInUse": {
    #             "enabledServices": enabledServices,
    #             "activeServices": activeServices,
    #             "package": installedPackage,
    #         }
    #     }
    # )

def generic_duplicate_entries(filename: str, separator: str, column: int) -> tuple:
    if column is not None and separator is not None:
        return {
            "ensureNoDuplicateEntriesExist": {
                "filename": filename,
                "separator": separator,
                "column": column,
            }
        }

    return {
        "ensureNoDuplicateEntriesExist": {
            "filename": filename,
        }
    }

def generic_ensure_no_duplicate_entries(row, match, action: str) -> tuple:
    if(match.group(1) == "user names"):
        return (
            None,
            generic_duplicate_entries("/etc/passwd", ":", 0)
        )
    elif(match.group(1) == "group names"):
        return (
            None,
            generic_duplicate_entries("/etc/group", ":", 0)
        )
    elif(match.group(1) == "UIDs"):
        return (
            None,
            generic_duplicate_entries("/etc/passwd", ":", 2)
        )
    elif(match.group(1) == "GIDs"):
        return (
            None,
            generic_duplicate_entries("/etc/group", ":", 2)
        )

    return None

def generic_ensure_sshd_configuration(row, match, action: str) -> tuple:
    procedure = load_procedure(row, AUDIT_IDX)
    prefix = '# sshd -T'
    if not procedure.startswith(prefix):
        print(f"Error: Invalid procedure for {row[RECOMMENDATION_IDX]} - {row[SUMMARY_IDX]}")
        print(procedure)
        return None

    print("Prefix matched")
    begin = procedure.find('\n\n')
    if begin == -1:
        print(f"Error: Invalid procedure for {row[RECOMMENDATION_IDX]} - {row[SUMMARY_IDX]}")
        print(procedure)
        return None
    begin += 2

    print("Blank line found")
    procedure = procedure[begin:]
    end = procedure.find('\n')
    if end != -1:
        procedure = procedure[:end]
    print(f"Procedure: {procedure}")
    procedure_parts = procedure.split()
    if len(procedure_parts) != 2:
        print(f"Error: Invalid procedure format for {row[RECOMMENDATION_IDX]} - {row[SUMMARY_IDX]}")
        print(procedure)
        return None

    return (
        None,
        {
            "ensureSshdConfigurationMatches": {
                "option": procedure_parts[0],
                "value": procedure_parts[1],
            }
        }
    )

    pass

known_patterns = {
    r"Ensure permissions on (/[^ ]+) are configured": generic_ensure_file_permissions,
    r"Ensure (\w+) option set on (/[^ ]+) partition": generic_ensure_partition_options1,
    r"Ensure (/[^ ]+) partition includes the (\w+) option": generic_ensure_partition_options2,
    r"Ensure mounting of (\w+) filesystems is disabled": generic_ensure_mounting_filesystems_disabled,
    r"Ensure (\w+) server services are not in use": generic_ensure_service_not_in_use,
    r"Ensure no duplicate ([\w\s]+) exist": generic_ensure_no_duplicate_entries,
    r"Ensure sshd (\w+) is disabled": generic_ensure_sshd_configuration,
}

def load_script(row, procedure_idx) -> str:
    if not row[procedure_idx]:
        return None

    # Extract text between ``` delimiters in the procedure field
    procedure = row[procedure_idx]
    if not procedure.startswith('Run the following script'):
        return None

    # print(procedure)
    script_start = procedure.find('```') + 3
    script_end = procedure.find('```', script_start)
    if script_start == -1 or script_end == -1:
        return None
    return row[procedure_idx][script_start:script_end].strip()

def load_generic_init(row):
    for pattern, callback in known_patterns.items():
        match = re.match(pattern, row[SUMMARY_IDX])
        if match:
            result = callback(row, match, 'init')
            if result:
                return (
                    result[0],
                    {
                        "init": [ result[1] ]
                    }
                )

    return (
        None,
        {
            "fail": {
                "message": "Not implemented yet"
            }
        }
    )

def load_generic_audit(row):
    for pattern, callback in known_patterns.items():
        match = re.match(pattern, row[SUMMARY_IDX])
        if match:
            result = callback(row, match, 'audit')
            if result:
                return (
                    result[0],
                    {
                        "audit": [ result[1] ]
                    }
                )

    # Failed to match a known pattern - fallback to bash script if available
    script = load_script(row, AUDIT_IDX)
    if script:
        func = { "bash": { "contents": script } }
        return (
            None,
            {
                "audit": [ func ]
            }
        )

    return (
        None,
        {
            "fail": {
                "message": "Not implemented yet"
            }
        }
    )

def load_generic_remediation(row) -> tuple:
    for pattern, callback in known_patterns.items():
        match = re.match(pattern, row[SUMMARY_IDX])
        if match:
            result = callback(row, match, 'remediation')
            if result:
                return (
                    result[0],
                    {
                        "remediate": [ result[1] ]
                    }
                )

    # Failed to match a known pattern - fallback to bash script if available
    script = load_script(row, REMEDIATION_IDX)
    if script:
        func = { "bash": { "contents": script } }
        return (
            None,
            {
                "remediate": [ func ]
            }
        )

    return (
        None,
        {
            "fail": {
                "message": "Not implemented yet"
            }
        }
    )

def load_recommendation(row) -> Recommendation:
    print(f"Loading recommendation: {row[RECOMMENDATION_IDX]} - {row[SUMMARY_IDX]}")
    init = load_generic_init(row)
    audit = load_generic_audit(row)
    remediation = load_generic_remediation(row)
    return Recommendation(
        row[RECOMMENDATION_IDX],
        row[SUMMARY_IDX],
        row[DESCRIPTION_IDX],
        # To be filled in, just placeholders for now
        init[1],
        audit[1],
        remediation[1],
        [audit[0]]
    )

def load_benchmark(sheet, level: CISLevel, machine_type: MachineType, distribution: str, distribution_version: str, version: str) -> CISBenchmark:
    benchmark = CISBenchmark(distribution, distribution_version, version, level.name, machine_type.name)

    print(f"Loading benchmark: {benchmark.distribution} {benchmark.distribution_version} {benchmark.version} {benchmark.level} {benchmark.machine_type}")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Skip rows without recommendation number
        if row[RECOMMENDATION_IDX] is None:
            continue
        # Skip rows without 'Automated' assessment status
        if row[ASSESMENT_STATUS_IDX] != 'Automated':
            continue

        # Load a single recommendation and store in the result
        benchmark.recommendations.append(load_recommendation(row))

    return benchmark

def parse_CIS_Excel(input_directory: str) -> CISModel:
    if not path.isdir(input_directory):
        print(f"Error: {input_directory} must be a directory.")
        sys.exit(1)

    model = CISModel()
    for input_definition in input_definitions.get_input_definitions():
        if input_definition.benchmark_name != "cis":
            print(f"Error: Unsupported benchmark name: {input_definition.benchmark_name}")
            sys.exit(1)

        print(f"Processing file: {path.join(input_directory, input_definition.file_name)}")

        # List all sheet names in the workbook
        workbook = openpyxl.load_workbook(path.join(input_directory, input_definition.file_name))

        if 'Level 1 - Server' in workbook.sheetnames:
            model.benchmarks.append(load_benchmark(workbook["Level 1 - Server"], CISLevel.L1, MachineType.Server, input_definition.distribution_name, input_definition.distribution_version, input_definition.benchmark_version))

        if 'Level 1 - Workstation' in workbook.sheetnames:
            model.benchmarks.append(load_benchmark(workbook["Level 1 - Workstation"], CISLevel.L1, MachineType.Workstation, input_definition.distribution_name, input_definition.distribution_version, input_definition.benchmark_version))

        if 'Level 2 - Server' in workbook.sheetnames:
            model.benchmarks.append(load_benchmark(workbook["Level 2 - Server"], CISLevel.L2, MachineType.Server, input_definition.distribution_name, input_definition.distribution_version, input_definition.benchmark_version))

        if 'Level 2 - Workstation' in workbook.sheetnames:
            model.benchmarks.append(load_benchmark(workbook["Level 2 - Workstation"], CISLevel.L1, MachineType.Workstation, input_definition.distribution_name, input_definition.distribution_version, input_definition.benchmark_version))
    return model
